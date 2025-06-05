#!/usr/bin/env python
# Beebe_Recall_postprocess.py

"""
NAME            Beebe_Recall_postprocess.py
AUTHOR          Brian Gravitt
VERSION         1
LAST UPDATE     2017-02-17
REQUIREMENTS    Python 3.x (https://www.python.org/)
                    (This script will not work in Python 2.x)
                The input files in the expected location/format
DESCRIPTION     This script is designed to parse text files from the Beebe client.
                It performs Beebee translations on the contents of the files and creates a new
                text file.  It then copies the input files to the newbiz directory.
HISTORY         Version 1.0 on 2017-02-17 - initial script creation
 
"""


# standard imports
import csv
import pyodbc
import datetime
import glob
import logging
import os
import shutil
import sys
import time
import traceback
from sys import argv
from collections import defaultdict
import xlsxwriter

# third party imports

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Alignment, Protection, Font

# setup logging
os.chdir("..")
# go up one directory
loggingdir = os.path.abspath(os.curdir) # capture the directory
logging_time = time.strftime("%Y%m%d%H%M%S")
logging_filename = ' Python Error Log.txt'
logging_filename_full = f'{loggingdir}/{logging_time}{logging_filename}'

def getKey(item):
    return item[1]

# main function
def main():

    ########################
    # setup and validation #
    ########################
    
    # directory setup #
    ###################
    
    # user feedback
    print()
    print("Verifying that all folders and files are accounted for...")
    print()
    os.chdir("..")
    #os.chdir("..")
    # create the variable for the Input directory (we already went back one directory for logging)
    inputdir = os.path.abspath(os.curdir) 
    # create the variable for the Output directory (go back one directory to the main Beebe folder)
    outputdir = os.path.abspath(os.curdir)
    # create the variable for the Archive directory (Beebe)
    archivedir = os.path.abspath(os.curdir) + '\\Data Processing\\EO - Archive\\Recall'
    
    # verify that the inputdir is readable
    try:
        os.access(inputdir, os.R_OK)
    except:
        sys.exit("The directory " + inputdir + " is not readable!  Check for valid folder or check folder permissions!  Terminating!")
    # verify that the outputdir is actually writable
    try:
        os.access(outputdir, os.W_OK)
    except:
        sys.exit("The directory " + outputdir + " is not writable!  Check for valid folder or check folder permissions!  Terminating!")
        
    # user feedback
    print("Input Directory: " + inputdir)
    print("Output Directory: " + outputdir)
    print("Archive Directory: " + archivedir)
    print()
    
    ###################
    # get input files #
    ###################
    
    #script, filedate = argv    
    #print(inputdir + "\\HCA_Recall_"  + filedate + "_1.txt")

    text_file_list = []
    items_to_keep = ["HCA_Recall_"]
  
    # iterate through any txt files in the input directory
    for file in glob.glob(inputdir + "\\" + "*.txt"):
        # capture the file if it contains any of our desired items
        if any(items in file for items in items_to_keep):
            text_file_list.append(file)
            
    # terminate if there are no text files
    if len(text_file_list) == 0:
        sys.exit("None of our desired files are in the directory!  Terminating...")
    
    # user feedback
    print("Text Files Located: " + str(text_file_list))
    print()
    

    ########################
    # process output files #
    ########################
    
    # user feedback
    print("Writing the Top of Output files...")
    print()
    
    # define the date for use in filenames
    today_as_mmddyy = datetime.datetime.now().strftime('%m%d%y')
    filedate = str(today_as_mmddyy)
    today_as_mmddyyslash = datetime.datetime.now().strftime('%m/%d/%y')

    #######################
    # process input files #
    #######################
    ncc = pyodbc.connect('DSN=HCA;UID=_system;PWD=SYS')
    cursor = ncc.cursor()

    # user feedback
    print("Processing the input files...")
    print()
    
        # create lists to hold lines and records
    all_lines = []
    ack = []
    bypass_due_to_ins = []
    could_not_recall = []

    # initial variable allocation
    r = 0
    rb = 0.
    i = 0
    ib = 0.
    c = 0
    cb = 0.
    clientlist = []
    clodatestr = ''
    clodatenow = datetime.datetime.now()

    # iterate through captured files
    for file in text_file_list:
        # open the file
        with open(file, "r") as f:
            # iterate over the file lines
            reader = list(csv.reader(f))
            for row in reader:
                if row:
                    if row[0] == 'ACK':
                        print(str(row[1]))
                        facs = """
                            SELECT ds.Disposition,
                                ds.Cancel_Reason,
                                dcgi.Client,
                                dcgi.AcctNUM_From_Client,
                                dcgi.Service_Date,
                                cg.CLIENT_NAME,
                                cg.Address_Line_1,
                                ds.ACCOUNT_NUM,
                                dgi.Last_Name,
                                dgi.First_Name,
                                db.Account_Balance,
                                dcgi.Date_Listed,
                                ds.Amount_Canceled,
                                dbcr.Cli_File_Recall
                            FROM SQLUser.Dbtr_Status ds
                            Join SQLUser.Dbtr_Clnt_Generl_Inf dcgi
                            on dcgi.ACCOUNT_NUM = ds.ACCOUNT_NUM
                            Join SQLUser.Dbtr_General_Inf dgi
                            on dgi.ACCOUNT_NUM = ds.ACCOUNT_NUM
                            Join SQLUser.Dbtr_Balances db
                            on db.ACCOUNT_NUM = ds.ACCOUNT_NUM
                            Join SQLUser.Clnt_General cg
                            on cg.CLIENT_NUM = dcgi.Client
                            Join SQLUser.DU_BEEBE_CLOSE_REQUESTS dbcr
                            on dbcr.ACCOUNT_NUM = ds.ACCOUNT_NUM
                            Where ds.ACCOUNT_NUM1 = '%(td)s'
                        """ % { 'td' : str(row[1])}

                        try:
                            cursor.execute(facs)
                        except pyodbc.Error as err:
                            sys.exit(err)
                            debug_msg(debug_level, 1, 'SQL', 'true')

                        #Print to Ack Regardless
                        #'Client Number','Division','NCC Account Num','Client Account Number','Service Date','Name'
                        for db in cursor.fetchall():
                            acct = str(db.ACCOUNT_NUM)
                            t = db.Service_Date
                            l = db.Date_Listed
                            balance = str(float("{0:.2f}".format(db.Account_Balance + db.Amount_Canceled)))
                            ack.append(db.Client + '|' +  db.CLIENT_NAME +  '|' + acct + '|' +  db.AcctNUM_From_Client  + '|' + t.strftime('%m/%d/%y') + '|' + l.strftime('%m/%d/%y') + '|' + db.First_Name + ' '+ db.Last_Name + '|' + balance)
                            c += 1
                            cb = cb + float(balance)
                            #Now Prin Exceptions
                            disp = str(db.Disposition)
                            try:
                               clodatenow  = db.Cli_File_Recall
                               clodatestr = clodatenow.strftime('%Y-%m-%d')
                            except:
                                clodatestr = ''

                            if clodatestr == '':
                                could_not_recall.append(db.Disposition + '|' + db.Client + '|'+ db.CLIENT_NAME + '|' + acct + '|' + db.AcctNUM_From_Client + '|' + t.strftime('%m/%d/%y') + '|'  +  l.strftime('%m/%d/%y') + '|' + db.First_Name + ' '+ db.Last_Name + '|' + balance)
                                r += 1
                                rb = rb + float(balance)
                            else:
                                if disp.startswith('9') == False:
                                    if disp.startswith('27') == True:
                                        bypass_due_to_ins.append(db.Disposition + '|' + db.Client + '|'+ db.CLIENT_NAME + '|' + acct + '|' + db.AcctNUM_From_Client + '|' + t.strftime('%m/%d/%y') + '|' + l.strftime('%m/%d/%y') + '|' + db.First_Name + ' '+ db.Last_Name + '|' + balance)
                                        i += 1
                                        ib = ib + float(balance)
                                    else:
                                        if disp.startswith('3PRC') == False:
                                            could_not_recall.append(db.Disposition + '|' + db.Client + '|'+ db.CLIENT_NAME + '|' + acct + '|' + db.AcctNUM_From_Client + '|' + t.strftime('%m/%d/%y') + '|'  +  l.strftime('%m/%d/%y') + '|' + db.First_Name + ' '+ db.Last_Name + '|' + balance)
                                            r += 1
                                            rb = rb + float(balance)
                            
                        
    # user feedback
    print("Input files have been processed!")
    print()
 
            
    ########################
    # process output files #
    ########################

    cursor.close
    # user feedback
    print("Writing the output files...")
    print()
    
    # define the date for use in filenames
    today_as_mmddyy = datetime.datetime.now().strftime('%m%d%y')
    
              
    filename = outputdir + '\\Beebe_Recall_Ack' + filedate + '.xlsx'
     
    workbook = xlsxwriter.Workbook(filename)
  
    # create the Excel formats
    date_format = workbook.add_format({'num_format': 'm/d/yyy', 'align': 'right'})
    month_format = workbook.add_format({'num_format': 'MMM-YY', 'align': 'right'})
    amt_format = workbook.add_format({'num_format': '_($* #,##0.00_);_($* (#,##0.00));_($* "-"??_);_(@_)', 'align': 'right'})
    total_format = workbook.add_format({'num_format': '#,###,##0.00', 'align': 'right', 'bold': True, 'bg_color': 'yellow'})
    header = workbook.add_format({'bold': True, 'bg_color': 'green', 'font_size': 26, 'align': 'center'})
    box14 = workbook.add_format({'bold': True, 'border': 1, 'font_size': 14,})
    box = workbook.add_format({'bold': True, 'border': 1,})
    highlight = workbook.add_format({'bold': True, 'bg_color': 'yellow'})
    percentage = workbook.add_format({'num_format': '%0.00', 'align': 'right'})
    total_percentage = workbook.add_format({'num_format': '%0.00', 'bold': True, 'bg_color': 'yellow'})
    net_format = workbook.add_format({'bold': True, 'underline': 'single'})
               
     ##############################################
    # create the excel spreadsheet for this file #
    ##############################################
          
 
    wb = Workbook()
    ws = wb.active
    ws.title = 'Recall Acknowledgments'

    ws['A1'] = 'Recall Acknowledgments'
    ws['A2'] = ''
    ws['A3'] = 'Processed on [' + time.strftime('%Y-%m-%d') + '] at [' + time.strftime('%I:%M %p') + ']'
            
        # define the headers
    ws_headers = ['Client Number','Division','NCC Account Num','Client Account Number','Service Date','List Date','Name','Balance']
            
    # write headers
    x = 5
    y = 1
    for h in ws_headers:
        c = ws.cell(row = x, column = y)
        c.value = h

        y = y + 1
    # set column dimensions
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 15

    # freeze panes
    ws.freeze_panes = ws['A6']
    fbalance = 0.
    acktot = 0
    ackbal = 0.

    clitot = 0 
    clibal = 0.

    prevcli = ''
    cliname = ''

    # write body
    x = 6
    y = 1
    prevrow = ""
    ack.sort()
    for row in ack:
        parts = row.split('|')
        if row != prevrow:
            for col in parts:
                c = ws.cell(row = x, column = y)
                if y == 1:
                    #Check Client
                    if prevcli == '':
                        prevcli = col
                        clitot += 1
                    else:
                        if prevcli == col:
                            clitot += 1
                        else:
                            c = ws.cell(row = x, column = y )
                            c.value = "Totals for " + cliname

                            c = ws.cell(row = x, column = y + 1)

                            c = ws.cell(row = x, column = y + 2)
                            c.value = clitot

                            c = ws.cell(row = x, column = y + 7)
                            c.value = clibal

                            acktot += clitot
                            ackbal += clibal
                            clitot = 1
                            clibal = 0
                            x += 1
                            c = ws.cell(row = x, column = y)
                            prevcli = col
                if y == 2:
                    cliname = col

                if y == 8:
                    fbalance = float(col)
                    c.value = fbalance

                    clibal += fbalance
                else:
                    c.value = col

                y += 1
            prevrow = row
            x += 1
        y = 1

    c = ws.cell(row = x, column = y)
    c.value = "Totals for " + cliname

    c = ws.cell(row = x, column = y + 1)

    c = ws.cell(row = x, column = y + 2)
    c.value = clitot

    c = ws.cell(row = x, column = y + 7)
    c.value = clibal

    x += 1
    acktot += clitot
    ackbal += clibal


    c = ws.cell(row = x, column = y)
    c.value = " Grand Totals"

    c = ws.cell(row = x, column = y + 2)
    c.value = acktot

    c = ws.cell(row = x, column = y + 7)
    c.value = ackbal

    # save the workbook to the designated path and filename
    wb.save(outputdir + '\\Beebe_Recall_Ack' + filedate + '.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Recall Exceptions'

    ws['A1'] = 'Recall Exceptions'
    ws['A2'] = 'From the file [' + file + ']'
    ws['A3'] = 'Processed on [' + time.strftime('%Y-%m-%d') + '] at [' + time.strftime('%I:%M %p') + ']'
            
        # define the headers
    ws_headers = ['Disposition','Client Number','Division','NCC Account Num','Client Account Number','Service Date', 'List Date', 'Name','Balance']
            
    # write headers
    x = 5
    y = 1
    for h in ws_headers:
        c = ws.cell(row = x, column = y)
        c.value = h

        y = y + 1
    # set column dimensions
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 15

    # freeze panes
    ws.freeze_panes = ws['A6']
                
    # write body
    x = 6
    y = 1
    prevrow = ""
    could_not_recall.sort()
    for row in could_not_recall:
        parts = row.split('|')
        if row != prevrow:
            for col in parts:
                c = ws.cell(row = x, column = y)
                c.value = col
                if y == 9:
                    c.value = float(col)

                else:
                    c.value = col
 
                y = y + 1
            prevrow = row
            x = x + 1
        y = 1

        # save the workbook to the designated path and filename
    wb.save(outputdir + '\\Beebe_Recall_Exp' + filedate + '.xlsx')
    
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Recalls in Insurance'

    ws['A1'] = 'Recall Exceptions Due to Insurance'
    ws['A2'] = 'From the file [' + file + ']'
    ws['A3'] = 'Processed on [' + time.strftime('%Y-%m-%d') + '] at [' + time.strftime('%I:%M %p') + ']'
            
        # define the headers
    ws_headers = ['Disposition','Client Number','Division','NCC Account Num','Client Account Number','Service Date', 'List Date', 'Name','Balance']
            
    # write headers
    x = 5
    y = 1
    for h in ws_headers:
        c = ws.cell(row = x, column = y)
        c.value = h

        y = y + 1
    # set column dimensions
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 15

    # freeze panes
    ws.freeze_panes = ws['A6']
                
    # write body
    x = 6
    y = 1
    prevrow = ""
    bypass_due_to_ins.sort()
    for row in bypass_due_to_ins:
        parts = row.split('|')
        if row != prevrow:
            for col in parts:
                c = ws.cell(row = x, column = y)
                c.value = col
                if y == 9:
                    c.value = float(col)

                else:
                    c.value = col

                y = y + 1
            prevrow = row
            x = x + 1
        y = 1

        # save the workbook to the designated path and filename
    wb.save(outputdir + '\\Beebe_Recall_Ins' + filedate + '.xlsx')    
    # user feedback
    print("Output files have been written!")
    print()
            
    #########################################
    # move input files to the newbiz folder #
    #########################################
    
    # user feedback
    print("Moving the input files to the Recall folder...")
    print()
    
    #for file in text_file_list:
    #    filename = os.path.basename(file) # get filename from full path
    #    shutil.move(file, archivedir + '\\' + filename)
        
    # user feedback
    print("Input files have been moved!")
    print()
    print("Goodbye!")
    print()
    
if __name__ == "__main__":
    try:
        main()
    except:
        traceback.print_exc(file=open(logging_filename_full,"a"))
        sys.exit(1)